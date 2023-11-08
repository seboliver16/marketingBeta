
import openpyxl
import datetime
import pandas as pd
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from bs4 import BeautifulSoup
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
import requests
import re
import os
import time


def parse_follower_count(count_str):
    if any(letter.isalpha() for letter in count_str):
        return 1000000  # Assume more than 100,000 if there's a letter
    return int(count_str)


#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Jun 12 15:20:22 2023

@author: seboliver
"""


def filter_human_accounts(followers_list):
    business_words = [
                    'shop', 'store', 'official', 'brand', 'marketing', 'services', 'consult',
                    'biz', 'company', 'inc', 'essay', 'writing', 'prof', 'photo',
                    'write', 'clean', 'solution', 'assign', 'help', 'world', 'kitchen', 'village', 
                    'sublease', 'roommate', 'stepmom', 'contact', 'nonstop', 'prep',
                    'poverty', 'grade', 'service', 'academic', 'homework',
                    'llc', 'bio', 'bios', 'tutor', 'tutoring', 'meet', 'class', 'order', 
                    'team', 'estate', 'club', 'org', 'nonprofit', 'project',
                    'yik', 'yak', 'barstool', 'camp', 'free', 'study', 'abroad', 'family', 
                    'student', 'discount', 'delivery', 'global', 'trade', 'tech', 
                    'solution', 'innovate', 'digital', 'network', 'finance', 
                    'investment', 'realtor', 'sales', 'deal', 'coupon', 
                    'bargain', 'enterprise', 'strategy', 'product', 'online', 
                    'virtual', 'cloud', 'app', 'platform', 'media', 
                    'advert', 'promo', 'ebook', 'download', 'subscribe', 
                    'membership', 'affiliate', 'partner', 'sponsor', 'webinar', 
                    'seminar', 'coach', 'mentor', 'guru', 'expert', 
                    'consultant', 'strategy', 'leader', 'agency', 'premium', 
                    'luxury', 'elite', 'pro', 'international', 'wholesale', 
                    'retail', 'commerce', 'commercial', 'industrial', 'mechanical',
                    'automation', 'bot', 'robot', 'AI', 'algorithm', 
                    'data', 'analytics', 'SEO', 'optimization', 'traffic', 
                    'click', 'subscribe', 'follower', 'influencer', 'monetize', 
                    'payment', 'credit', 'loan', 'finance', 'insurance', 
                    'legal', 'attorney', 'lawyer', 'claim', 'compensation', 
                    'pharma', 'med', 'clinic', 'surgery', 'pill', 
                    'supplement', 'diet', 'fitness', 'gym', 'workout', 
                    'beauty', 'cosmetic', 'fashion', 'style', 'trend', 
                    'vlog', 'blog', 'content', 'post', 'publish', 
                    'editor', 'journal', 'news', 'headline', 'press', 
                    'release', 'statement', 'announcement', 'update', 'info', 
                    'review', 'feedback', 'testimonial', 'rating', 'vote', 
                    'survey', 'quiz', 'question', 'inquiry', 'form', 
                    'report', 'download', 'access', 'login', 'signup', 
                    'register', 'enroll', 'join', 'start', 'launch', 
                    'event', 'session', 'workshop', 'course', 'lesson', 'jonahliss',
                    'module', 'tutorial', 'guide', 'manual', 'book', 
                    'read', 'learn', 'educate', 'train', 'develop', 
                    'skill', 'talent', 'career', 'job', 'employment', 
                    'hire', 'staff', 'team', 'crew', 'associate', 
                    'colleague', 'partner', 'collaborate', 'cooperate', 'network', 
                    'connect', 'link', 'contact', 'email', 'message', 
                    'chat', 'talk', 'speak', 'call', 'phone', 
                    'video', 'camera', 'photo', 'image', 'pic', 
                    'audio', 'sound', 'music', 'song', 'track', 
                    'game', 'play', 'fun', 'entertainment', 'show', 
                    'watch', 'view', 'stream', 'broadcast', 'live', 
                    'host', 'guest', 'audience', 'fan', 'supporter', 
                    'follower', 'friend', 'member', 'community', 'group', 
                    'forum', 'board', 'committee', 'council', 'leadership', 
                    'president', 'director', 'manager', 'executive', 'officer', 
                    'secretary', 'clerk', 'assistant', 'helper', 'volunteer', 
                    'donor', 'contributor', 'support', 'help', 'service', 
                    'benefit', 'advantage', 'feature', 'option', 'choice', 
                    'offer', 'deal', 'sale', 'discount', 'bargain', 
                    'price', 'cost', 'fee', 'charge', 'rate', 
                    'budget', 'expense', 'payment', 'invoice', 'transaction', 
                    'order', 'purchase', 'buy', 'shop', 'store', 
                    'cart', 'basket', 'delivery', 'ship', 'mail', 
                    'address', 'location', 'place', 'site', 'area', 
                    'space', 'property', 'building', 'construction', 'architecture', 
                    'design', 'plan', 'strategy', 'approach', 'method', 
                    'process', 'system', 'program', 'project', 'task', 
                    'work', 'job', 'duty', 'responsibility', 'obligation', 
                    'requirement', 'rule', 'regulation', 'policy', 'guideline', 
                    'standard', 'quality', 'excellence', 'perfection', 'achievement', 
                    'success', 'win', 'victory', 'award', 'prize', 
                    'trophy', 'medal', 'certificate', 'degree', 'diploma', 
                    'education', 'school', 'college', 'university', 'academy', 
                    'institute', 'center', 'clinic', 'hospital', 'pharmacy', 
                    'restaurant', 'cafe', 'bar', 'pub', 'club', 
                    'hotel', 'resort', 'vacation', 'travel', 'tour',
                    'shop', 'store', 'official', 'brand', 'marketing', 'services', 'consult'
                    'biz', 'company', 'inc', 'essay', 'writing', 'prof', 'photo',
                    'write', 'clean', 'solution', 'assign', 'help', 'world', 'kitchen', 'village', 'sublease', 'roommate', 'stepmom', 'contact', 'nonstop', 'prep'
                    'poverty', 'grade', 'service', 'academic', 'grade', 'academic', 'homework',
                    'llc', 'bio', 'bios', 'tutor', 'tutoring', 'meet', 'class', 'order', 'team', 'estate', 'club', 'org', 'nonprofit', 'project', '2027','2028'
                    'yik', 'yak', 'barstool', 'camp', 'free', 'study', 'abroad', 'family', 'student']
    business_pattern = re.compile('|'.join(business_words), re.IGNORECASE)
    return [follower for follower in followers_list if not business_pattern.search(follower)]


def send_discord_message(webhook_url, message):
    data = {
        "content": message
    }
    requests.post(webhook_url, json=data)


def scrape(user, password, scrapeUser):

    try:
        # Send start message

        # Configure the web driver (replace 'chromedriver' with the path to your driver executable)
        driver = webdriver.Chrome()

        driver.get('https://www.instagram.com/accounts/login/')
        wait = WebDriverWait(driver, 10)
        time.sleep(2)

        # Log in
        username_field = driver.find_element(
            By.CSS_SELECTOR, "input[name='username']")
        password_field = driver.find_element(
            By.CSS_SELECTOR, "input[name='password']")
        username_field.send_keys(user)
        password_field.send_keys(password)
        password_field.send_keys(Keys.RETURN)
        time.sleep(10)

        # Navigate to the account
        driver.get(f'https://www.instagram.com/{scrapeUser}/')

        followers_count_element = wait.until(EC.presence_of_element_located(
            (By.CSS_SELECTOR, 'a[href$="/followers/"]')))
        
        followers_count = int(
            followers_count_element.text.replace(',', '').replace('K', '000').replace('M', '000000').replace('.', '').split(' ')[0])
        limit = int(followers_count * 0.65)
        if limit > 2200:
            limit = 2200
        print("Follower Count is ", followers_count)
        # Open the followers list
        followers_count_element.click()
        time.sleep(5)

        # Scroll the followers list
        fBody = driver.find_element(
            By.CSS_SELECTOR, 'div._aano')
        followers = set()
        print(f"Follower count for {scrapeUser} is ", str(followers_count))
        # Capture the start time
        start_time = datetime.datetime.now()

        while True:
            html = driver.page_source

            # Calculate the elapsed time
            elapsed_time = (datetime.datetime.now() - start_time).total_seconds()

            # Break the loop if elapsed time is more than 60 seconds
            if elapsed_time > 600:
                print("Exiting loop after 10 mins")
                break

            driver.execute_script(
                'arguments[0].scrollTop = arguments[0].scrollTop + arguments[0].offsetHeight;', fBody)

            time.sleep(1)
            soup = BeautifulSoup(html, 'html.parser')
            followers_tmp = soup.findAll('span', class_='_ap3a _aaco _aacw _aacx _aad7 _aade')
            follow_buttons = driver.find_elements(By.CSS_SELECTOR, 'button._acan')  # Locate all "Follow" buttons
            
            for idx, follower in enumerate(followers_tmp):
                
                try:
                    button_text = follow_buttons[idx].text  # Try to get the text from the corresponding "Follow" button
                    if(button_text == "Follow"):
                        followers.add(follower.text)
                    
                except IndexError:
                    pass
            
            print(len(followers))
            if len(followers) >= limit:
                break
        # Filter followers and create DataFrame
        # Filter followers and create DataFrame
        filtered_followers = filter_human_accounts(list(followers))
        filename = f'{user}.xlsx'

        # Old Code
        # if os.path.exists(filename):
        #     book = openpyxl.load_workbook(filename)
        #     sheet = book.active

        
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = "Followers"
        headers = ["Follower", "Contacted"]
        sheet.append(headers)

        # existing_followers = [row[0] for row in sheet.iter_rows(
        #     min_row=2, max_row=sheet.max_row, max_col=1, values_only=True)]


        for follower in filtered_followers:
            # if follower not in existing_followers:
            row_data = [follower, False]
            sheet.append(row_data)

        book.save(filename)

    except Exception as e:
        # Send error message
        error_message = f"Error: {str(e)}"
        print(f"ERROR in createFollowerList: {e}")